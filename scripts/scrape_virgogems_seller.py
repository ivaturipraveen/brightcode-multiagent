"""
Generalized Virgo Gems scraper — pass seller slug as argument.
Usage: python3 scrape_virgogems_seller.py mangoman
"""
import sys
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import time
import json
from datetime import datetime

SELLER = sys.argv[1] if len(sys.argv) > 1 else "mangoman"
BASE_URL = f"https://virgogems.com/?s={SELLER}&v=0b3b97fa6688"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36"
}


def get_total_pages(soup, base_url):
    nav = soup.find("nav", class_=re.compile(r"pagination|nav-links", re.I))
    if nav:
        nums = []
        for a in nav.find_all("a"):
            try:
                nums.append(int(a.text.strip()))
            except:
                pass
        if nums:
            return max(nums)
    title = soup.find("title")
    if title:
        m = re.search(r"of\s+(\d+)", title.text)
        if m:
            return int(m.group(1))
    # Probe page 2
    probe = scrape_page(base_url + "&paged=2")
    if probe:
        t = probe.find("title")
        if t:
            m = re.search(r"of\s+(\d+)", t.text)
            if m:
                return int(m.group(1))
    return 1


def parse_items_from_text(text):
    items = []
    pattern = re.compile(
        r'([A-Za-z][A-Za-z\s/\-()]+?)\-\s*([A-Za-z][A-Za-z\s,\.]+?)\s+(\d+(?:\.\d+)?)',
        re.MULTILINE
    )
    for match in pattern.finditer(text):
        mineral = match.group(1).strip()
        location = match.group(2).strip()
        price = match.group(3).strip()
        if len(mineral) > 1 and len(location) > 1 and not mineral.lower().startswith("shipping"):
            items.append({"mineral": mineral, "location": location, "price_usd": float(price)})
    return items


def parse_shipping_from_text(text):
    m = re.search(r'[Ss]hipping\s+(\d+)', text)
    return float(m.group(1)) if m else None


def scrape_page(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None


def scrape_all(seller, base_url):
    print(f"[{seller}] Fetching page 1...")
    soup = scrape_page(base_url)
    if not soup:
        return []

    total_pages = get_total_pages(soup, base_url)
    print(f"[{seller}] Total pages: {total_pages}")

    all_listings = []
    listing_id = 1

    def process_soup(soup, page_num):
        nonlocal listing_id
        articles = soup.find_all(["article", "li"], class_=re.compile(r"post|product|result", re.I))
        if not articles:
            articles = soup.find_all(["p", "div"], class_=re.compile(r"entry-content|post-content|summary", re.I))

        if articles:
            for article in articles:
                text = article.get_text(separator=" ", strip=True)
                if len(text) < 10:
                    continue
                title_tag = article.find(["h1","h2","h3","a"])
                title = title_tag.get_text(strip=True) if title_tag else ""
                a_tag = article.find("a", href=True)
                link = a_tag["href"] if a_tag else ""
                date_tag = article.find(["time","span"], class_=re.compile(r"date|time|published", re.I))
                post_date = date_tag.get_text(strip=True) if date_tag else ""
                img_tag = article.find("img")
                img_url = img_tag.get("src","") if img_tag else ""
                items = parse_items_from_text(text)
                shipping = parse_shipping_from_text(text)
                if items:
                    for item in items:
                        all_listings.append({
                            "ID": listing_id, "Page": page_num,
                            "Post Title": title, "Post URL": link,
                            "Post Date": post_date, "Post Image URL": img_url,
                            "Mineral / Item": item["mineral"],
                            "Origin / Location": item["location"],
                            "Price (USD)": item["price_usd"],
                            "Shipping (USD)": shipping if shipping else "",
                            "Total w/ Shipping": round(item["price_usd"] + shipping, 2) if shipping else "",
                            "Raw Text Snippet": text[:300],
                        })
                        listing_id += 1
                else:
                    all_listings.append({
                        "ID": listing_id, "Page": page_num,
                        "Post Title": title, "Post URL": link,
                        "Post Date": post_date, "Post Image URL": img_url,
                        "Mineral / Item": "", "Origin / Location": "",
                        "Price (USD)": "", "Shipping (USD)": shipping if shipping else "",
                        "Total w/ Shipping": "", "Raw Text Snippet": text[:300],
                    })
                    listing_id += 1
        else:
            body = soup.find("body")
            if body:
                text = body.get_text(separator=" ", strip=True)
                items = parse_items_from_text(text)
                shipping = parse_shipping_from_text(text)
                for item in items:
                    all_listings.append({
                        "ID": listing_id, "Page": page_num,
                        "Post Title": f"Page {page_num} listing",
                        "Post URL": base_url + f"&paged={page_num}",
                        "Post Date": "", "Post Image URL": "",
                        "Mineral / Item": item["mineral"],
                        "Origin / Location": item["location"],
                        "Price (USD)": item["price_usd"],
                        "Shipping (USD)": shipping if shipping else "",
                        "Total w/ Shipping": round(item["price_usd"] + shipping, 2) if shipping else "",
                        "Raw Text Snippet": text[:300],
                    })
                    listing_id += 1

    process_soup(soup, 1)
    for page in range(2, total_pages + 1):
        url = f"{base_url}&paged={page}"
        print(f"[{seller}] Fetching page {page}/{total_pages}...")
        s = scrape_page(url)
        if s:
            process_soup(s, page)
        time.sleep(0.4)

    return all_listings


def write_excel(listings, path, seller):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"

    hdr_fill = PatternFill("solid", fgColor="003882")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EEF3FB")
    border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:L1")
    ws["A1"] = f"Virgo Gems — {seller}  |  Scraped: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  Total records: {len(listings)}"
    ws["A1"].font = Font(bold=True, size=12, color="003882")
    ws["A1"].fill = PatternFill("solid", fgColor="D0E4FF")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    headers = ["ID","Page #","Post Title","Post URL","Post Date","Post Image URL",
               "Mineral / Item","Origin / Location","Price (USD)","Shipping (USD)","Total w/ Shipping","Raw Text Snippet"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[2].height = 20

    for row_idx, item in enumerate(listings, 3):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        vals = [item["ID"], item["Page"], item["Post Title"], item["Post URL"],
                item["Post Date"], item["Post Image URL"], item["Mineral / Item"],
                item["Origin / Location"], item["Price (USD)"], item["Shipping (USD)"],
                item["Total w/ Shipping"], item["Raw Text Snippet"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if row_fill: cell.fill = row_fill
            cell.alignment = left if col > 2 else center
            cell.border = border
        ws.row_dimensions[row_idx].height = 18

    col_widths = [6,7,30,45,16,45,28,22,12,14,16,50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(listings)+2}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary Statistics"
    ws2["A1"].font = Font(bold=True, size=13, color="003882")
    prices = [r["Price (USD)"] for r in listings if isinstance(r["Price (USD)"], (int,float))]
    origins = {}; minerals = {}
    for r in listings:
        if r["Origin / Location"]:
            origins[r["Origin / Location"]] = origins.get(r["Origin / Location"], 0) + 1
        if r["Mineral / Item"]:
            minerals[r["Mineral / Item"]] = minerals.get(r["Mineral / Item"], 0) + 1
    summary = [
        ("Total Records", len(listings)),
        ("Total Pages Scraped", max((r["Page"] for r in listings), default=0)),
        ("Items with Prices", len(prices)),
        ("Min Price (USD)", min(prices) if prices else "N/A"),
        ("Max Price (USD)", max(prices) if prices else "N/A"),
        ("Avg Price (USD)", round(sum(prices)/len(prices),2) if prices else "N/A"),
        ("Total Value (USD)", round(sum(prices),2) if prices else "N/A"),
        ("Unique Origins", len(origins)),
        ("Unique Minerals", len(minerals)),
        ("Scraped At (UTC)", datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
    ]
    for r, (k,v) in enumerate(summary, 3):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20
    ws2["A14"] = "Top 10 Origins"; ws2["A14"].font = Font(bold=True, color="003882")
    for r, (o,c) in enumerate(sorted(origins.items(), key=lambda x:-x[1])[:10], 15):
        ws2.cell(row=r, column=1, value=o); ws2.cell(row=r, column=2, value=c)
    ws2["D14"] = "Top 10 Minerals"; ws2["D14"].font = Font(bold=True, color="003882")
    for r, (m,c) in enumerate(sorted(minerals.items(), key=lambda x:-x[1])[:10], 15):
        ws2.cell(row=r, column=4, value=m); ws2.cell(row=r, column=5, value=c)

    wb.save(path)
    return {
        "total": len(listings), "prices": len(prices),
        "min_price": min(prices) if prices else 0,
        "max_price": max(prices) if prices else 0,
        "avg_price": round(sum(prices)/len(prices),2) if prices else 0,
        "total_value": round(sum(prices),2) if prices else 0,
        "unique_origins": len(origins), "unique_minerals": len(minerals),
        "pages": max((r["Page"] for r in listings), default=0),
    }


if __name__ == "__main__":
    listings = scrape_all(SELLER, BASE_URL)
    print(f"\n[{SELLER}] Total records: {len(listings)}")
    out = f"/home/ubuntu/openclaw-multiagent/virgo_gems_{SELLER}.xlsx"
    stats = write_excel(listings, out, SELLER)
    print(json.dumps(stats))
    # Write stats to a sidecar json for parent to read
    with open(f"/tmp/virgo_{SELLER}_stats.json", "w") as f:
        json.dump(stats, f)
    print(f"[{SELLER}] Saved to {out}")
