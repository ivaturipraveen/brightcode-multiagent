"""
Scrape all search results from virgogems.com for @the_mystic.raven
and export to Excel with all available data points.
"""
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import time
import json
from datetime import datetime

BASE_URL = "https://virgogems.com/?s=%40the_mystic.raven&v=0b3b97fa6688"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36"
}

def get_total_pages(soup):
    # Check for pagination
    nav = soup.find("nav", class_=re.compile(r"pagination|nav-links", re.I))
    if nav:
        page_nums = nav.find_all("a")
        nums = []
        for a in page_nums:
            try:
                nums.append(int(a.text.strip()))
            except:
                pass
        if nums:
            return max(nums)
    # Try title like "Page 2 of 30"
    title = soup.find("title")
    if title:
        m = re.search(r"of\s+(\d+)", title.text)
        if m:
            return int(m.group(1))
    return 1

def parse_items_from_text(text):
    """Parse mineral/gem entries from raw listing text."""
    items = []
    # Split on entries like "MineralName- Location Price"
    # Pattern: Name(s)- Location Price (number)
    pattern = re.compile(
        r'([A-Za-z][A-Za-z\s/\-()]+?)\-\s*([A-Za-z][A-Za-z\s,\.]+?)\s+(\d+(?:\.\d+)?)',
        re.MULTILINE
    )
    for match in pattern.finditer(text):
        mineral = match.group(1).strip()
        location = match.group(2).strip()
        price = match.group(3).strip()
        # Filter out noise
        if len(mineral) > 1 and len(location) > 1 and not mineral.lower().startswith("shipping"):
            items.append({
                "mineral": mineral,
                "location": location,
                "price_usd": float(price),
            })
    return items

def parse_shipping_from_text(text):
    m = re.search(r'[Ss]hipping\s+(\d+)', text)
    return float(m.group(1)) if m else None

def scrape_page(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None

def scrape_all():
    print("Fetching page 1...")
    soup = scrape_page(BASE_URL)
    if not soup:
        return []

    total_pages = get_total_pages(soup)
    # Fallback: we know from live inspection this search has 30 pages
    if total_pages == 1:
        # Probe page 2 to see if it exists and check title
        probe = scrape_page(BASE_URL + "&paged=2")
        if probe:
            t = probe.find("title")
            if t:
                m = re.search(r"of\s+(\d+)", t.text)
                if m:
                    total_pages = int(m.group(1))
    print(f"Total pages: {total_pages}")

    all_listings = []
    listing_id = 1

    def process_soup(soup, page_num):
        nonlocal listing_id
        # Try to find WooCommerce products or blog posts
        articles = soup.find_all(["article", "li"], class_=re.compile(r"post|product|result", re.I))
        if not articles:
            # Fallback: grab all <p> or content divs
            articles = soup.find_all(["p", "div"], class_=re.compile(r"entry-content|post-content|summary", re.I))

        if articles:
            for article in articles:
                text = article.get_text(separator=" ", strip=True)
                if len(text) < 10:
                    continue

                # Get title/link
                title_tag = article.find(["h1","h2","h3","a"])
                title = title_tag.get_text(strip=True) if title_tag else ""
                link = ""
                a_tag = article.find("a", href=True)
                if a_tag:
                    link = a_tag["href"]

                # Get date if available
                date_tag = article.find(["time", "span"], class_=re.compile(r"date|time|published", re.I))
                post_date = date_tag.get_text(strip=True) if date_tag else ""

                # Get image
                img_tag = article.find("img")
                img_url = img_tag.get("src","") if img_tag else ""

                # Parse items
                items = parse_items_from_text(text)
                shipping = parse_shipping_from_text(text)

                if items:
                    for item in items:
                        all_listings.append({
                            "ID": listing_id,
                            "Page": page_num,
                            "Post Title": title,
                            "Post URL": link,
                            "Post Date": post_date,
                            "Post Image URL": img_url,
                            "Mineral / Item": item["mineral"],
                            "Origin / Location": item["location"],
                            "Price (USD)": item["price_usd"],
                            "Shipping (USD)": shipping if shipping else "",
                            "Total w/ Shipping": round(item["price_usd"] + shipping, 2) if shipping else "",
                            "Raw Text Snippet": text[:300],
                        })
                        listing_id += 1
                else:
                    # Still record the post even if no items parsed
                    all_listings.append({
                        "ID": listing_id,
                        "Page": page_num,
                        "Post Title": title,
                        "Post URL": link,
                        "Post Date": post_date,
                        "Post Image URL": img_url,
                        "Mineral / Item": "",
                        "Origin / Location": "",
                        "Price (USD)": "",
                        "Shipping (USD)": shipping if shipping else "",
                        "Total w/ Shipping": "",
                        "Raw Text Snippet": text[:300],
                    })
                    listing_id += 1
        else:
            # Last resort: grab body text
            body = soup.find("body")
            if body:
                text = body.get_text(separator=" ", strip=True)
                items = parse_items_from_text(text)
                shipping = parse_shipping_from_text(text)
                for item in items:
                    all_listings.append({
                        "ID": listing_id,
                        "Page": page_num,
                        "Post Title": f"Page {page_num} listing",
                        "Post URL": BASE_URL + f"&paged={page_num}",
                        "Post Date": "",
                        "Post Image URL": "",
                        "Mineral / Item": item["mineral"],
                        "Origin / Location": item["location"],
                        "Price (USD)": item["price_usd"],
                        "Shipping (USD)": shipping if shipping else "",
                        "Total w/ Shipping": round(item["price_usd"] + shipping, 2) if shipping else "",
                        "Raw Text Snippet": text[:300],
                    })
                    listing_id += 1

    process_soup(soup, 1)

    for page in range(2, total_pages + 1):
        url = f"https://virgogems.com/?s=%40the_mystic.raven&v=0b3b97fa6688&paged={page}"
        print(f"Fetching page {page}/{total_pages}...")
        s = scrape_page(url)
        if s:
            process_soup(s, page)
        time.sleep(0.4)  # polite crawl

    return all_listings

def write_excel(listings, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Virgo Gems Listings"

    # ── Styles ──
    hdr_fill = PatternFill("solid", fgColor="003882")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EEF3FB")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Meta banner ──
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Virgo Gems — Search Results for @the_mystic.raven  |  Scraped: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  Total records: {len(listings)}"
    ws["A1"].font = Font(bold=True, size=12, color="003882")
    ws["A1"].fill = PatternFill("solid", fgColor="D0E4FF")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    # ── Headers ──
    headers = [
        "ID", "Page #", "Post Title", "Post URL", "Post Date",
        "Post Image URL", "Mineral / Item", "Origin / Location",
        "Price (USD)", "Shipping (USD)", "Total w/ Shipping", "Raw Text Snippet"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    # ── Data rows ──
    for row_idx, item in enumerate(listings, 3):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        vals = [
            item["ID"], item["Page"], item["Post Title"], item["Post URL"],
            item["Post Date"], item["Post Image URL"], item["Mineral / Item"],
            item["Origin / Location"], item["Price (USD)"], item["Shipping (USD)"],
            item["Total w/ Shipping"], item["Raw Text Snippet"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left if col > 2 else center
            cell.border = border
        ws.row_dimensions[row_idx].height = 18

    # ── Column widths ──
    col_widths = [6, 7, 30, 45, 16, 45, 28, 22, 12, 14, 16, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze header rows ──
    ws.freeze_panes = "A3"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(listings)+2}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary Statistics"
    ws2["A1"].font = Font(bold=True, size=13, color="003882")

    prices = [r["Price (USD)"] for r in listings if isinstance(r["Price (USD)"], (int, float))]
    origins = {}
    minerals = {}
    for r in listings:
        if r["Origin / Location"]:
            origins[r["Origin / Location"]] = origins.get(r["Origin / Location"], 0) + 1
        if r["Mineral / Item"]:
            minerals[r["Mineral / Item"]] = minerals.get(r["Mineral / Item"], 0) + 1

    summary = [
        ("Total Records", len(listings)),
        ("Total Pages Scraped", max((r["Page"] for r in listings), default=0)),
        ("Items with Prices", len(prices)),
        ("Min Price (USD)", min(prices) if prices else "N/A"),
        ("Max Price (USD)", max(prices) if prices else "N/A"),
        ("Avg Price (USD)", round(sum(prices)/len(prices), 2) if prices else "N/A"),
        ("Total Value (USD)", round(sum(prices), 2) if prices else "N/A"),
        ("Unique Origins", len(origins)),
        ("Unique Minerals", len(minerals)),
        ("Scraped At (UTC)", datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
    ]
    for r, (k, v) in enumerate(summary, 3):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    # Top origins
    ws2["A14"] = "Top 10 Origins"
    ws2["A14"].font = Font(bold=True, color="003882")
    for r, (origin, count) in enumerate(sorted(origins.items(), key=lambda x: -x[1])[:10], 15):
        ws2.cell(row=r, column=1, value=origin)
        ws2.cell(row=r, column=2, value=count)

    # Top minerals
    ws2["D14"] = "Top 10 Minerals"
    ws2["D14"].font = Font(bold=True, color="003882")
    for r, (mineral, count) in enumerate(sorted(minerals.items(), key=lambda x: -x[1])[:10], 15):
        ws2.cell(row=r, column=4, value=mineral)
        ws2.cell(row=r, column=5, value=count)

    wb.save(path)
    print(f"Saved: {path}")
    return {
        "total": len(listings),
        "prices": len(prices),
        "min_price": min(prices) if prices else 0,
        "max_price": max(prices) if prices else 0,
        "avg_price": round(sum(prices)/len(prices), 2) if prices else 0,
        "total_value": round(sum(prices), 2) if prices else 0,
        "unique_origins": len(origins),
        "unique_minerals": len(minerals),
    }

if __name__ == "__main__":
    listings = scrape_all()
    print(f"\nTotal records scraped: {len(listings)}")
    out_path = "/home/ubuntu/openclaw-multiagent/virgo_gems_listings.xlsx"
    stats = write_excel(listings, out_path)
    print(json.dumps(stats, indent=2))
